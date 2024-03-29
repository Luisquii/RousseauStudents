from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl as xl
import datetime
from PIL import Image, ImageTk
import webbrowser
from openpyxl.styles import Font
import os
import shutil

class Rousseau:
    global alumnosxlsxFlag
    global rowToEditedStudent
    global names_list


    def __init__(self, master):
        self.initUI(master)
        self.findStudent = None
        self.updateStudent = None
        self.editStudent = None


    #Windows
    def initUI(self, master):

        self.master = master

        master.title("Rousseau")

        # Labels
        usernameLabel = tk.Label(master, text="Username")
        #        self.usernameLabel.pack(side=LEFT, padx=1, pady=1)
        usernameLabel.grid(row=0)

        passwordLabel = tk.Label(master, text="Password")
        passwordLabel.grid(row=1)

        # Entries
        self.usernameEntry = tk.Entry(master)
        self.usernameEntry.grid(row=0, column=1)
        self.usernameEntry.focus()

        self.passwordEntry = tk.Entry(master, show="*")
        self.passwordEntry.grid(row=1, column=1)

        # Buttons
        okButton = tk.Button(master, text="Entrar", fg="black", command=self.getAndVerifyUsernameAndPassword).grid(
            row=2, column=1, sticky=E)

    def mainMenuUI(self):
        def goToWeb(url):
            webbrowser.open_new(url)
        self.mainMenu = Toplevel(self.master)
        self.mainMenu.iconbitmap("imgs/Papalote.ico")
        self.mainMenu.title("Menu Principal")
        self.mainMenu.geometry("350x150+50+100")
        self.mainMenu.resizable(False,False)

        # Buttons & Binds
        newStudentButton = tk.Button(self.mainMenu, text="Nuevo Alumno", width=20, command=self.newStudentUI)
        newStudentButton.place(x = 0, y = 20)

        findStudentButton = tk.Button(self.mainMenu, text="Buscar Alumno", width=20, command=self.findStudentUI)
        findStudentButton.place(x = 0, y = 50)

        updateStudentButton = tk.Button(self.mainMenu, text = "Actualizar Alumnos", width = 20, command = self.updateStudentUI)
        updateStudentButton.place(x=0, y = 80)

        #Hyperlink
        webpageLink = tk.Label(self.mainMenu, text = "Colegio Rousseau Webpage", fg = "blue", cursor = "hand2")
        webpageLink.place(x = 50, y = 130)
        webpageLink.bind("<Button-1>", lambda e: goToWeb("http://colegio-rousseau.edu.mx/"))

        # Images
        load = Image.open("imgs\LogoBlanco.png")
        load = load.resize((125, 125), Image.ANTIALIAS)
        render = ImageTk.PhotoImage(load)
        img = Label(self.mainMenu, image=render)
        img.image = render
        img.place(x=200, y=5)

    def newStudentUI(self):
        self.newStudent = Toplevel(self.mainMenu)
        self.newStudent.iconbitmap("imgs/Papalote.ico")
        self.newStudent.title("Agregar nuevo alumno")
        self.newStudent.geometry("1175x700")

        # Labels
        nombreLabel = tk.Label(self.newStudent, text="Apellidos y Nombres:").grid(row=0, column=0, sticky=E)
        fechaNacimientoLabel = tk.Label(self.newStudent, text="Fecha de nacimiento:").grid(row=0, column=2, sticky=E)
        curpLabel = tk.Label(self.newStudent, text="Curp:").grid(row=0, column=4, sticky=E)

        gradoACursarLabel = tk.Label(self.newStudent, text="Grado a cursar:").grid(row=1, column=0, sticky=E)
        cicloEscolarLabel = tk.Label(self.newStudent, text="Ciclo escolar:").grid(row=1, column=2, sticky=E)
        escuelaProcedenciaLabel = tk.Label(self.newStudent, text="Escuela de procedencia:").grid(row=1, column=4)
        claveLabel = tk.Label(self.newStudent, text="Clave:").grid(row=1, column=6, sticky=E)
        conQuienViveLabel = tk.Label(self.newStudent, text="Con quien vive:").grid(row=3, column=0, sticky=E)

        #############
        datosGeneralesLabel = tk.Label(self.newStudent, text="Datos Generales").grid(row=5, column=0, sticky=E)
        calleLabel = tk.Label(self.newStudent, text="Calle y Numero:").grid(row=6, column=0, sticky=E)
        coloniaLabel = tk.Label(self.newStudent, text="Colonia:").grid(row=6, column=2, sticky=E)
        entreCalles = tk.Label(self.newStudent, text="Entre que calles:").grid(row=6, column=4, sticky=E)

        cpLabel = tk.Label(self.newStudent, text="C.P.:").grid(row=7, column=0, sticky=E)
        ciudadLabel = tk.Label(self.newStudent, text="Ciudad:").grid(row=7, column=2, sticky=E)
        telLabel = tk.Label(self.newStudent, text="Tel:").grid(row=7, column=4, sticky=E)
        tel2Label = tk.Label(self.newStudent, text="Otro Tel:").grid(row=7, column=6, sticky=E)

        religionLabel = tk.Label(self.newStudent, text="Religion:").grid(row=8, column=0, sticky=E)
        enfermedadesLabel = tk.Label(self.newStudent, text="Enfermedades o accidentes:").grid(row=8, column=2, sticky=E)
        sangreLabel = tk.Label(self.newStudent, text="T. sangre").grid(row=8, column=6, sticky=E)

        tratmientoLabel = tk.Label(self.newStudent, text="Actualmente en tratamiento:").grid(row=9, column=0, sticky=E,
                                                                                             columnspan=2)
        servicioMedicoLabel = tk.Label(self.newStudent, text="Servicio medico al que pertenece:").grid(row=10, column=0,
                                                                                                       sticky=E,
                                                                                                       columnspan=2)
        otroLabel = tk.Label(self.newStudent, text="Otro:").grid(row=10, column=5, sticky=E)

        ############PoT
        datosImportantesLabel = tk.Label(self.newStudent,
                                         text="DATOS IMPORTANTES PARA CAPTURAR AL SISTEMA INTEGRAL DE INFORMACION EDUCATIVA DEL MODULO DE CONTROL ESCOLAR DE LA S.E.P.").grid(
            row=12, column=0, columnspan=6)

        nombrePoTLabel = tk.Label(self.newStudent, text="Nombre padre o tutor:").grid(row=13, column=0, sticky=E)
        fechaNacimientoPoTLabel = tk.Label(self.newStudent, text="Fecha de nacimiento:").grid(row=13, column=2,
                                                                                              sticky=E)
        curpPoTLabel = tk.Label(self.newStudent, text="Curp:").grid(row=13, column=4, sticky=E)

        rfcPoTLabel = tk.Label(self.newStudent, text="R.F.C:").grid(row=14, column=0, sticky=E)
        lugarNacimientoPoTLabel = tk.Label(self.newStudent, text="Lugar de nacimiento:").grid(row=14, column=2,
                                                                                              sticky=E)
        estadoCivilPoTLabel = tk.Label(self.newStudent, text="Estado Civil:").grid(row=14, column=4, sticky=E)

        nacionalidadPoTLabel = tk.Label(self.newStudent, text="Nacionalidad:").grid(row=15, column=0, sticky=E)
        profesionPoTLabel = tk.Label(self.newStudent, text="Profesion:").grid(row=15, column=2, sticky=E)
        telefonoPoTLabel = tk.Label(self.newStudent, text="Telefono:").grid(row=15, column=4, sticky=E)

        celularPoTLabel = tk.Label(self.newStudent, text="Celular:").grid(row=16, column=0, sticky=E)
        lugarTrabajoPoTLabel = tk.Label(self.newStudent, text="Lugar de trabajo:").grid(row=16, column=2, sticky=E)
        ocupacionPoTLabel = tk.Label(self.newStudent, text="Ocupacion:").grid(row=16, column=4, sticky=E)

        emailPoTLabel = tk.Label(self.newStudent, text="E-Mail:").grid(row=17, column=0, sticky=E)

        ###########MoT
        guionMoTLabel = tk.Label(self.newStudent, text=" ------------------------------------------------"
                                                       "------------------------------------------------"
                                                       "------------------------------------------------").grid(row=18,
                                                                                                                column=1,
                                                                                                                columnspan=5)

        nombreMoTLabel = tk.Label(self.newStudent, text="Nombre madre o tutora:").grid(row=19, column=0, sticky=E)
        fechaNacimientoMoTLabel = tk.Label(self.newStudent, text="Fecha de nacimiento:").grid(row=19, column=2,
                                                                                              sticky=E)
        curpMoTLabel = tk.Label(self.newStudent, text="Curp:").grid(row=19, column=4, sticky=E)

        rfcMoTLabel = tk.Label(self.newStudent, text="R.F.C:").grid(row=20, column=0, sticky=E)
        lugarNacimientoMoTLabel = tk.Label(self.newStudent, text="Lugar de nacimiento:").grid(row=20, column=2,
                                                                                              sticky=E)
        estadoCivilMoTLabel = tk.Label(self.newStudent, text="Estado Civil:").grid(row=20, column=4, sticky=E)

        nacionalidadMoTLabel = tk.Label(self.newStudent, text="Nacionalidad:").grid(row=21, column=0, sticky=E)
        profesionMoTLabel = tk.Label(self.newStudent, text="Profesion:").grid(row=21, column=2, sticky=E)
        telefonoMoTLabel = tk.Label(self.newStudent, text="Telefono:").grid(row=21, column=4, sticky=E)

        celularMoTLabel = tk.Label(self.newStudent, text="Celular:").grid(row=22, column=0, sticky=E)
        lugarTrabajoMoTLabel = tk.Label(self.newStudent, text="Lugar de trabajo:").grid(row=22, column=2, sticky=E)
        ocupacionMoTLabel = tk.Label(self.newStudent, text="Ocupacion:").grid(row=22, column=4, sticky=E)

        emailMoTLabel = tk.Label(self.newStudent, text="E-Mail:").grid(row=23, column=0, sticky=E)

        referenciaLabel = tk.Label(self.newStudent, text="RECIBIO REFERENCIA DE NUESTRA INSTITUCION A TRAVES DE:").grid(
            row=26, column=1, columnspan=5)

        siguienteAnoLabel = tk.Label(self.newStudent, text = "MARQUE LA CASILLA SI EL ALUMNO A INSCRIBIR EMPEZARA EL SIGUIENTE AÑO:", fg="green").grid(row = 28, column = 3, columnspan = 5, sticky = W)

        datosCompletosLabel = tk.Label(self.newStudent,
                                       text="***REVISAR QUE TODOS LOS DATOS ESTEN COMPLETOS Y CORRECTOS***", fg="red").grid(
            row=29, column=1, columnspan=4)

        # Choices
        # Create a Tkinter variable
        self.cicloEscolarVar = StringVar(self.newStudent)
        cicloEscolarChoices = ["2019-2020", "2020-2021", "2021-2022", "2022-2023", "2023-2024", "2024-2025",
                               "2025-2026", "2026-2027", "2027-2028", "2028-2029", "2029-2030"]
        self.cicloEscolarVar.set('2019-2020')  # set the default option
        cicloEscolarpopupMenu = tk.OptionMenu(self.newStudent, self.cicloEscolarVar, *cicloEscolarChoices).grid(row=1,
                                                                                                                column=3,
                                                                                                                sticky=W)

        self.gradoACursarVar = StringVar(self.newStudent)
        gradoACursarChoices = ["1 Kinder", "2 Kinder", "3 Kinder", "1 Primaria", "2 Primaria", "3 Primaria",
                               "4 Primaria", "5 Primaria", "6 Primaria", "1 Secundaria", "2 Secundaria", "3 Secundaria"]
        self.gradoACursarVar.set("1 Kinder")  # Set the default option
        gradoACursarpopupMenu = tk.OptionMenu(self.newStudent, self.gradoACursarVar, *gradoACursarChoices).grid(row=1,
                                                                                                                column=1,
                                                                                                                sticky=W)

        # Entries
        self.nombreEntry = tk.Entry(self.newStudent, width=30);
        self.nombreEntry.grid(row=0, column=1);
        self.nombreEntry.focus()
        self.fechaNacimientoEntry = tk.Entry(self.newStudent);
        self.fechaNacimientoEntry.grid(row=0, column=3, sticky=W)
        self.curpEntry = tk.Entry(self.newStudent);
        self.curpEntry.grid(row=0, column=5, sticky=W)

        self.escuelaProcedenciaEntry = tk.Entry(self.newStudent, width=25);
        self.escuelaProcedenciaEntry.grid(row=1, column=5)
        self.claveEntry = tk.Entry(self.newStudent, width=10);
        self.claveEntry.grid(row=1, column=7, sticky=W)

        self.calleEntry = tk.Entry(self.newStudent, width=30);
        self.calleEntry.grid(row=6, column=1, sticky=W)
        self.coloniaEntry = tk.Entry(self.newStudent, width=25);
        self.coloniaEntry.grid(row=6, column=3, sticky=W)
        self.entreCallesEntry = tk.Entry(self.newStudent, width=49);
        self.entreCallesEntry.grid(row=6, column=5, sticky=W, columnspan=3)

        self.cpEntry = tk.Entry(self.newStudent);
        self.cpEntry.grid(row=7, column=1, sticky=W)
        self.ciudadEntry = tk.Entry(self.newStudent);
        self.ciudadEntry.grid(row=7, column=3, sticky=W)
        self.telEntry = tk.Entry(self.newStudent);
        self.telEntry.grid(row=7, column=5, sticky=W)
        self.tel2Entry = tk.Entry(self.newStudent);
        self.tel2Entry.grid(row=7, column=7, sticky=W)

        self.religionEntry = tk.Entry(self.newStudent);
        self.religionEntry.grid(row=8, column=1, sticky=W)
        self.enfermedadesEntry = tk.Entry(self.newStudent, width=65);
        self.enfermedadesEntry.grid(row=8, column=3, sticky=W, columnspan=3)
        self.sangreEntry = tk.Entry(self.newStudent);
        self.sangreEntry.grid(row=8, column=7)

        self.otroEntry = tk.Entry(self.newStudent);
        self.otroEntry.grid(row=10, column=6)

        ############PoT
        self.nombrePoTEntry = tk.Entry(self.newStudent, width=30);
        self.nombrePoTEntry.grid(row=13, column=1, sticky=W)
        self.fechaNacimientoPoTEntry = tk.Entry(self.newStudent);
        self.fechaNacimientoPoTEntry.grid(row=13, column=3, sticky=W)
        self.curpPoTEntry = tk.Entry(self.newStudent);
        self.curpPoTEntry.grid(row=13, column=5, sticky=W)

        self.rfcPoTEntry = tk.Entry(self.newStudent);
        self.rfcPoTEntry.grid(row=14, column=1, sticky=W)
        self.lugarNacimientoPoTEntry = tk.Entry(self.newStudent);
        self.lugarNacimientoPoTEntry.grid(row=14, column=3, sticky=W)
        self.estadoCivilPoTEntry = tk.Entry(self.newStudent);
        self.estadoCivilPoTEntry.grid(row=14, column=5, sticky=W)

        self.nacionalidadPoTEntry = tk.Entry(self.newStudent);
        self.nacionalidadPoTEntry.grid(row=15, column=1, sticky=W)
        self.profesionPoTEntry = tk.Entry(self.newStudent);
        self.profesionPoTEntry.grid(row=15, column=3, sticky=W)
        self.telefonoPoTEntry = tk.Entry(self.newStudent);
        self.telefonoPoTEntry.grid(row=15, column=5, sticky=W)

        self.celularPoTEntry = tk.Entry(self.newStudent);
        self.celularPoTEntry.grid(row=16, column=1, sticky=W)
        self.lugarTrabajoPoTEntry = tk.Entry(self.newStudent);
        self.lugarTrabajoPoTEntry.grid(row=16, column=3, sticky=W)
        self.ocupacionPoTEntry = tk.Entry(self.newStudent);
        self.ocupacionPoTEntry.grid(row=16, column=5, sticky=W)

        self.emailPoTEntry = tk.Entry(self.newStudent);
        self.emailPoTEntry.grid(row=17, column=1, sticky=W)

        ###########MoT
        self.nombreMoTEntry = tk.Entry(self.newStudent, width=30);
        self.nombreMoTEntry.grid(row=19, column=1, sticky=W)
        self.fechaNacimientoMoTEntry = tk.Entry(self.newStudent);
        self.fechaNacimientoMoTEntry.grid(row=19, column=3, sticky=W)
        self.curpMoTEntry = tk.Entry(self.newStudent);
        self.curpMoTEntry.grid(row=19, column=5, sticky=W)

        self.rfcMoTEntry = tk.Entry(self.newStudent);
        self.rfcMoTEntry.grid(row=20, column=1, sticky=W)
        self.lugarNacimientoMoTEntry = tk.Entry(self.newStudent);
        self.lugarNacimientoMoTEntry.grid(row=20, column=3, sticky=W)
        self.estadoCivilMoTEntry = tk.Entry(self.newStudent);
        self.estadoCivilMoTEntry.grid(row=20, column=5, sticky=W)

        self.nacionalidadMoTEntry = tk.Entry(self.newStudent);
        self.nacionalidadMoTEntry.grid(row=21, column=1, sticky=W)
        self.profesionMoTEntry = tk.Entry(self.newStudent);
        self.profesionMoTEntry.grid(row=21, column=3, sticky=W)
        self.telefonoMoTEntry = tk.Entry(self.newStudent);
        self.telefonoMoTEntry.grid(row=21, column=5, sticky=W)

        self.celularMoTEntry = tk.Entry(self.newStudent);
        self.celularMoTEntry.grid(row=22, column=1, sticky=W)
        self.lugarTrabajoMoTEntry = tk.Entry(self.newStudent);
        self.lugarTrabajoMoTEntry.grid(row=22, column=3, sticky=W)
        self.ocupacionMoTEntry = tk.Entry(self.newStudent);
        self.ocupacionMoTEntry.grid(row=22, column=5, sticky=W)

        self.emailMoTEntry = tk.Entry(self.newStudent);
        self.emailMoTEntry.grid(row=23, column=1, sticky=W)

        # Checkboxes
        self.imssCBVar = BooleanVar();
        self.imssCBVar.set(False);
        self.isssteCBVar = BooleanVar();
        self.isssteCBVar.set(False);
        self.pemexCBVar = BooleanVar();
        self.pemexCBVar.set(False)
        imssCB = tk.Checkbutton(self.newStudent, text="IMSS", variable=self.imssCBVar).grid(row=10, column=2, sticky=E)
        issteCB = tk.Checkbutton(self.newStudent, text="ISSSTE", variable=self.isssteCBVar).grid(row=10, column=3,
                                                                                                 sticky=E)
        pemexCB = tk.Checkbutton(self.newStudent, text="PEMEX", variable=self.pemexCBVar).grid(row=10, column=4,
                                                                                               sticky=E)
        self.siguienteAnoCBVar = BooleanVar();
        siguienteAnoCB = tk.Checkbutton(self.newStudent, text = "", variable= self.siguienteAnoCBVar).grid(row = 28,column = 6, sticky = W)


        # RadioButtons
        self.responsabeRBVar = StringVar();
        self.responsabeRBVar.set(False)
        madreRB = tk.Radiobutton(self.newStudent, text="MADRE", variable=self.responsabeRBVar, value="Madre").grid(
            row=3, column=1)
        padreRB = tk.Radiobutton(self.newStudent, text="PADRE", variable=self.responsabeRBVar, value="Padre").grid(
            row=3, column=2)
        ambosRB = tk.Radiobutton(self.newStudent, text="AMBOS", variable=self.responsabeRBVar, value="Ambos").grid(
            row=3, column=3)
        tutorRB = tk.Radiobutton(self.newStudent, text="TUTOR(A)", variable=self.responsabeRBVar, value="Tutor").grid(
            row=3, column=4)

        self.sinoRBVar = StringVar();
        self.sinoRBVar.set(False)
        siRB = tk.Radiobutton(self.newStudent, text="Si", variable=self.sinoRBVar, value="Si").grid(row=9, column=2,
                                                                                                    sticky=E)
        noRB = tk.Radiobutton(self.newStudent, text="No", variable=self.sinoRBVar, value="No").grid(row=9, column=3,
                                                                                                    sticky=E)

        self.referenciaRBVar = StringVar();
        self.referenciaRBVar.set(False)
        directorioRB = tk.Radiobutton(self.newStudent, text="ANUNCIO DIRECTORIO", variable=self.referenciaRBVar,
                                      value="directorio").grid(row=27, column=1, sticky=E)
        periodicoRB = tk.Radiobutton(self.newStudent, text="REDES SOCIALES", variable=self.referenciaRBVar,
                                     value="periodico").grid(row=27, column=2, sticky=E)
        famoamistadRB = tk.Radiobutton(self.newStudent, text="FAMILIAR / AMISTAD", variable=self.referenciaRBVar,
                                       value="familiar/amistad").grid(row=27, column=3, sticky=E)
        webRB = tk.Radiobutton(self.newStudent, text="PAGINA WEB", variable=self.referenciaRBVar, value="Web").grid(
            row=27, column=4, sticky=E)
        espectacularRB = tk.Radiobutton(self.newStudent, text="ESPECTACULAR", variable=self.referenciaRBVar,
                                        value="espectacular").grid(row=27, column=5, sticky=E)

        # Buttons
        self.agregarButton = tk.Button(self.newStudent, text="Agregar", fg="black", command= self.getDataFromNewStudent)
        self.agregarButton.grid(row=29, column=5, sticky=E)
        self.limpiarButton = tk.Button(self.newStudent, text="Limpiar", fg="black",
                                       command=self.clearDataFromNewStudent)
        self.limpiarButton.grid(row=29, column=4, sticky=E)

        # ProgressBar
        self.progressBar = ttk.Progressbar(self.newStudent, length=100)
        self.progressBar.grid(row=30, column=6, sticky=E)

        # Images
        load = Image.open("imgs\LogoBlanco.png")
        load = load.resize((175, 175), Image.ANTIALIAS)
        render = ImageTk.PhotoImage(load)
        img = Label(self.newStudent, image=render)
        img.image = render
        img.place(x=0, y=515)

    def findStudentUI(self):
        nameSelected = None
        nameSelectedFlag = False

        def on_keyrelease(event):

            # get text from entry
            value = event.widget.get()
            value = value.strip().lower()

            # get data from names_list
            if value == '':
                data = self.names_list
            else:
                data = []
                for item in self.names_list:
                    if value in item.lower():
                        data.append(item)

            # update data in listbox
            listbox_update(data)

        def listbox_update(data):
            # delete previous data
            self.nombresListbox.delete(0, 'end')

            # sorting data
            data = sorted(data, key=str.lower)

            # put new data
            for item in data:
                self.nombresListbox.insert('end', item)

        def on_select(event):
            # Display and save the current name selected
            nonlocal nameSelected
            nonlocal nameSelectedFlag
            nameSelected = str(event.widget.get(event.widget.curselection()))
            nameSelectedFlag = True



        if not self.findStudent:
            self.findStudent = Toplevel(self.mainMenu)
            self.findStudent.iconbitmap("imgs/Papalote.ico")
            self.findStudent.title("Buscar Alumno")
            self.findStudent.geometry("1175x700")

        self.findStudent.protocol("WM_DELETE_WINDOW", self.destroyFindStudent)

        xlObj2 = rousseauXL()
        self.names_list = xlObj2.getNamesList()

        #Labels indicators
        self.nombreAlumnoLabel = tk.Label(self.findStudent, text = "Nombre del alumno:"); self.nombreAlumnoLabel.grid(column = 0, row = 0, sticky = W);
        self.nombreAlumnoSelectedLabel = tk.Label(self.findStudent, text = "Nombre del alumno:", font='Helvetica 10 bold'); self.nombreAlumnoSelectedLabel.grid(column = 0, row = 4, sticky = E);
        self.fechaNacimientoAlumnoLabel = tk.Label(self.findStudent, text = "Fecha de nacimiento:"); self.fechaNacimientoAlumnoLabel.grid(column = 0, row = 5, sticky = E);
        self.curpAlumnoLabel = tk.Label(self.findStudent, text= "CURP:");self.curpAlumnoLabel.grid(column=0, row=6, sticky=E);
        self.gradoACursarLabel = tk.Label(self.findStudent, text = "Grado en que fue inscrito:"); self.gradoACursarLabel.grid(column = 0, row = 7 , sticky = E);
        self.cicloEscolarLabel = tk.Label(self.findStudent, text = "Ciclo en que fue inscrito:"); self.cicloEscolarLabel.grid(column = 0, row = 8, sticky = E)
        self.gradoEnCursoLabel = tk.Label(self.findStudent, text = "Grado en curso:"); self.gradoEnCursoLabel.grid(column = 0, row = 9, sticky = E)
        self.cicloEnCursoLabel = tk.Label(self.findStudent, text = "Ciclo en Curso:"); self.cicloEnCursoLabel.grid(column = 0, row = 10, sticky = E)
        self.conQuienViveLabel = tk.Label(self.findStudent, text = "Con quien vive:"); self.conQuienViveLabel.grid(column = 0, row = 11, sticky = E)
        self.calleAlumnoLabel = tk.Label(self.findStudent, text = "Calle:"); self.calleAlumnoLabel.grid(column = 0 , row = 12, sticky = E)
        self.coloniaAlumnoLabel = tk.Label(self.findStudent, text = "Colonia:");self.coloniaAlumnoLabel.grid(column = 0, row = 13, sticky = E)
        self.cpAlumnoLabel = tk.Label(self.findStudent, text = "C.P.:"); self.cpAlumnoLabel.grid(column = 0, row = 14, sticky = E)
        self.ciudadAlumnoLabel = tk.Label(self.findStudent, text = "Ciudad:");self.ciudadAlumnoLabel.grid(column = 0, row = 15, sticky = E)
        self.entreCallesAlumnoLabel = tk.Label(self.findStudent, text = "Entre calles:"); self.entreCallesAlumnoLabel.grid(column = 0, row = 16, sticky = E)
        self.telefonoAlumnoLabel = tk.Label(self.findStudent, text = "Telefono:"); self.telefonoAlumnoLabel.grid(column = 0, row = 17, sticky = E)
        self.telefono2AlumnoLabel = tk.Label(self.findStudent, text = "Otro telefono:"); self.telefono2AlumnoLabel.grid(column = 0, row = 18, sticky = E)
        self.enfermedadesLabel = tk.Label(self.findStudent, text = "Enfermedades:"); self.enfermedadesLabel.grid(column = 0, row = 19, sticky = E)
        self.tipoSangreLabel = tk.Label(self.findStudent, text = "Tipo de sangre:"); self.tipoSangreLabel.grid(column = 0 , row = 20, sticky = E)
        self.enTratamientoLabel = tk.Label(self.findStudent, text = "En tratamiento:"); self.enTratamientoLabel.grid(column = 0, row = 21, sticky = E)
        self.servicioMedicoLabel = tk.Label(self.findStudent, text = "Servicio medico:"); self.servicioMedicoLabel.grid(column = 0, row = 22, sticky = E)
        self.fechaInscripcionLabel = tk.Label(self.findStudent, text = "Fecha de inscripcion:"); self.fechaInscripcionLabel.grid(column = 0, row = 23, sticky = E)
        self.statusLabel = tk.Label(self.findStudent, text = "Estado:"); self.statusLabel.grid(column = 0, row = 24, sticky = E)

        # PoT
        self.nombrePotLabel = tk.Label(self.findStudent, text = "Nombre del padre o tutor", font='Helvetica 10 bold'); self.nombrePotLabel.grid(column =2, row = 4, sticky = E)
        self.curpPotLabel= tk.Label(self.findStudent, text = "CURP:"); self.curpPotLabel.grid(column =2, row = 5, sticky = E)
        self.rfcPotLabel= tk.Label(self.findStudent, text = "RFC:"); self.rfcPotLabel.grid(column =2, row = 6, sticky = E)
        self.estadoCivilPotLabel= tk.Label(self.findStudent, text = "Estado Civil:"); self.estadoCivilPotLabel.grid(column =2, row = 7, sticky = E)
        self.nacionalidadPotLabel= tk.Label(self.findStudent, text = "Nacionalidad:"); self.nacionalidadPotLabel.grid(column =2, row = 8, sticky = E)
        self.profesionPotLabel= tk.Label(self.findStudent, text = "Profesion:"); self.profesionPotLabel.grid(column =2, row = 9, sticky = E)
        self.ocupacionPotLabel= tk.Label(self.findStudent, text = "Ocupacion:"); self.ocupacionPotLabel.grid(column =2, row = 10, sticky = E)
        self.lugarDeTrabajoPotLabel= tk.Label(self.findStudent, text = "Lugar de Trabajo:"); self.lugarDeTrabajoPotLabel.grid(column =2, row = 11, sticky = E)
        self.telefonoPotLabel= tk.Label(self.findStudent, text = "Telefono:"); self.telefonoPotLabel.grid(column =2, row = 12, sticky = E)
        self.celularPotLabel= tk.Label(self.findStudent, text = "Celular:"); self.celularPotLabel.grid(column =2, row = 13, sticky = E)
        self.emailPotLabel= tk.Label(self.findStudent, text = "Email:"); self.emailPotLabel.grid(column =2, row = 14, sticky = E)

        #MoT
        self.nombreMotLabel = tk.Label(self.findStudent, text = "Nombre del madre o tutora", font='Helvetica 10 bold'); self.nombreMotLabel.grid(column =4, row = 4, sticky = E)
        self.curpMotLabel= tk.Label(self.findStudent, text = "CURP:"); self.curpMotLabel.grid(column =4, row = 5, sticky = E)
        self.rfcMotLabel= tk.Label(self.findStudent, text = "RFC:"); self.rfcMotLabel.grid(column =4, row = 6, sticky = E)
        self.estadoCivilMotLabel= tk.Label(self.findStudent, text = "Estado Civil:"); self.estadoCivilMotLabel.grid(column =4, row = 7, sticky = E)
        self.nacionalidadMotLabel= tk.Label(self.findStudent, text = "Nacionalidad:"); self.nacionalidadMotLabel.grid(column =4, row = 8, sticky = E)
        self.profesionMotLabel= tk.Label(self.findStudent, text = "Profesion:"); self.profesionMotLabel.grid(column =4, row = 9, sticky = E)
        self.ocupacionMotLabel= tk.Label(self.findStudent, text = "Ocupacion:"); self.ocupacionMotLabel.grid(column =4, row = 10, sticky = E)
        self.lugarDeTrabajoMotLabel= tk.Label(self.findStudent, text = "Lugar de Trabajo:"); self.lugarDeTrabajoMotLabel.grid(column =4, row = 11, sticky = E)
        self.telefonoMotLabel= tk.Label(self.findStudent, text = "Telefono:"); self.telefonoMotLabel.grid(column =4, row = 12, sticky = E)
        self.celularMotLabel= tk.Label(self.findStudent, text = "Celular:"); self.celularMotLabel.grid(column =4, row = 13, sticky = E)
        self.emailMotLabel= tk.Label(self.findStudent, text = "Email:"); self.emailMotLabel.grid(column =4, row = 14, sticky = E)

        #Labels results
        self.nombreAlumnoResultLabel = tk.Label(self.findStudent, text = "-"); self.nombreAlumnoResultLabel.grid(column = 1, row = 4, sticky = W);
        self.fechaNacimientoAlumnoResultLabel = tk.Label(self.findStudent, text = "-"); self.fechaNacimientoAlumnoResultLabel.grid(column = 1, row = 5, sticky = W);
        self.curpAlumnoResultLabel= tk.Label(self.findStudent, text = "-"); self.curpAlumnoResultLabel.grid(column = 1, row = 6, sticky = W);
        self.gradoACursarResultLabel= tk.Label(self.findStudent, text = "-"); self.gradoACursarResultLabel.grid(column = 1, row = 7, sticky = W);
        self.cicloEscolarResultLabel= tk.Label(self.findStudent, text = "-"); self.cicloEscolarResultLabel.grid(column = 1, row = 8, sticky = W);
        self.gradoEnCursoResultLabel= tk.Label(self.findStudent, text = "-"); self.gradoEnCursoResultLabel.grid(column = 1, row = 9, sticky = W);
        self.cicloEnCursoResultLabel = tk.Label(self.findStudent, text = "-"); self.cicloEnCursoResultLabel.grid(column = 1, row = 10, sticky = W);

        self.conQuienViveResultLabel= tk.Label(self.findStudent, text = "-"); self.conQuienViveResultLabel.grid(column = 1, row = 11, sticky = W);
        self.calleAlumnoResultLabel= tk.Label(self.findStudent, text = "-"); self.calleAlumnoResultLabel.grid(column = 1, row = 12, sticky = W);
        self.coloniaAlumnoResultLabel= tk.Label(self.findStudent, text = "-"); self.coloniaAlumnoResultLabel.grid(column = 1, row = 13, sticky = W);
        self.cpAlumnoResultLabel= tk.Label(self.findStudent, text = "-"); self.cpAlumnoResultLabel.grid(column = 1, row = 14, sticky = W);
        self.ciudadAlumnoResultLabel= tk.Label(self.findStudent, text = "-"); self.ciudadAlumnoResultLabel.grid(column = 1, row = 15, sticky = W);
        self.entreCallesAlumnoResultLabel= tk.Label(self.findStudent, text = "-"); self.entreCallesAlumnoResultLabel.grid(column = 1, row = 16, sticky = W);
        self.telefonoAlumnoResultLabel= tk.Label(self.findStudent, text = "-"); self.telefonoAlumnoResultLabel.grid(column = 1, row = 17, sticky = W);
        self.telefono2AlumnoResultLabel= tk.Label(self.findStudent, text = "-"); self.telefono2AlumnoResultLabel.grid(column = 1, row = 18, sticky = W);
        self.enfermedadesResultLabel= tk.Label(self.findStudent, text = "-"); self.enfermedadesResultLabel.grid(column = 1, row = 19, sticky = W);
        self.tipoSangreResultLabel= tk.Label(self.findStudent, text = "-"); self.tipoSangreResultLabel.grid(column = 1, row = 20, sticky = W);
        self.enTratamientoResultLabel= tk.Label(self.findStudent, text = "-"); self.enTratamientoResultLabel.grid(column = 1, row = 21, sticky = W);
        self.servicioMedicoResultLabel= tk.Label(self.findStudent, text = "-"); self.servicioMedicoResultLabel.grid(column = 1, row = 22, sticky = W);
        self.fechaInscripcionResultLabel = tk.Label(self.findStudent, text = "-"); self.fechaInscripcionResultLabel.grid(column = 1, row = 23, sticky = W);
        self.statusResultLabel = tk.Label(self.findStudent, text = "-"); self.statusResultLabel.grid(column = 1, row = 24, sticky = W);

        # PoT
        self.nombrePotResultLabel = tk.Label(self.findStudent, text = "-"); self.nombrePotResultLabel.grid(column = 3, row = 4, sticky = W);
        self.curpPotResultLabel = tk.Label(self.findStudent, text = "-"); self.curpPotResultLabel.grid(column = 3, row = 5, sticky = W);
        self.rfcPotResultLabel = tk.Label(self.findStudent, text = "-"); self.rfcPotResultLabel.grid(column = 3, row = 6, sticky = W);
        self.estadoCivilPotResultLabel = tk.Label(self.findStudent, text = "-"); self.estadoCivilPotResultLabel.grid(column = 3, row = 7, sticky = W);
        self.nacionalidadPotResultLabel = tk.Label(self.findStudent, text = "-"); self.nacionalidadPotResultLabel.grid(column = 3, row = 8, sticky = W);
        self.profesionPotResultLabel = tk.Label(self.findStudent, text = "-"); self.profesionPotResultLabel.grid(column = 3, row = 9, sticky = W);
        self.ocupacionPotResultLabel = tk.Label(self.findStudent, text = "-"); self.ocupacionPotResultLabel.grid(column = 3, row = 10, sticky = W);
        self.lugarDeTrabajoPotResultLabel = tk.Label(self.findStudent, text = "-"); self.lugarDeTrabajoPotResultLabel.grid(column = 3, row = 11, sticky = W);
        self.telefonoPotResultLabel = tk.Label(self.findStudent, text = "-"); self.telefonoPotResultLabel.grid(column = 3, row = 12, sticky = W);
        self.celularPotResultLabel = tk.Label(self.findStudent, text = "-"); self.celularPotResultLabel.grid(column = 3, row = 13, sticky = W);
        self.emailPotResultLabel = tk.Label(self.findStudent, text = "-"); self.emailPotResultLabel.grid(column = 3, row = 14, sticky = W);

        # MoT
        self.nombreMotResultLabel  = tk.Label(self.findStudent, text = "-"); self.nombreMotResultLabel.grid(column = 6, row = 4, sticky = W);
        self.curpMotResultLabel  = tk.Label(self.findStudent, text = "-"); self.curpMotResultLabel.grid(column = 6, row = 5, sticky = W);
        self.rfcMotResultLabel  = tk.Label(self.findStudent, text = "-"); self.rfcMotResultLabel.grid(column = 6, row = 6, sticky = W);
        self.estadoCivilMotResultLabel  = tk.Label(self.findStudent, text = "-"); self.estadoCivilMotResultLabel.grid(column = 6, row = 7, sticky = W);
        self.nacionalidadMotResultLabel  = tk.Label(self.findStudent, text = "-"); self.nacionalidadMotResultLabel.grid(column = 6, row = 8, sticky = W);
        self.profesionMotResultLabel  = tk.Label(self.findStudent, text = "-"); self.profesionMotResultLabel.grid(column = 6, row = 9, sticky = W);
        self.ocupacionMotResultLabel  = tk.Label(self.findStudent, text = "-"); self.ocupacionMotResultLabel.grid(column = 6, row = 10, sticky = W);
        self.lugarDeTrabajoMotResultLabel  = tk.Label(self.findStudent, text = "-"); self.lugarDeTrabajoMotResultLabel.grid(column = 6, row = 11, sticky = W);
        self.telefonoMotResultLabel  = tk.Label(self.findStudent, text = "-"); self.telefonoMotResultLabel.grid(column = 6, row = 12, sticky = W);
        self.celularMotResultLabel  = tk.Label(self.findStudent, text = "-"); self.celularMotResultLabel.grid(column = 6, row = 13, sticky = W);
        self.emailMotResultLabel  = tk.Label(self.findStudent, text = "-"); self.emailMotResultLabel.grid(column = 6, row = 14, sticky = W);


        #Entries
        self.nombreAlumnoEntry = tk.Entry(self.findStudent, width = 60); self.nombreAlumnoEntry.grid(row = 1, column = 0); self.nombreAlumnoEntry.bind('<KeyRelease>', on_keyrelease)


        #Lists
        self.nombresListbox = tk.Listbox(self.findStudent, width=60); self.nombresListbox.grid(row = 2, column = 0);
        self.nombresListbox.bind('<<ListboxSelect>>', on_select)
        listbox_update(self.names_list)

        #Button
        self.buscarAlumnoButton = tk.Button(self.findStudent, text = "MOSTRAR INFORMACION", command= lambda: self.getDataFromSpecificStudent(xlObj2, nameSelected, nameSelectedFlag, False)); self.buscarAlumnoButton.grid(row = 3, column = 0, sticky = E)
        self.borrarAlumnoButton = tk.Button(self.findStudent, text = "BORRAR ALUMNO", command = lambda: self.deleteSpecificStudent(xlObj2, nameSelected, nameSelectedFlag)); self.borrarAlumnoButton.grid(row = 22, column = 8)
        self.editarAlumnoButton = tk.Button(self.findStudent, text = "EDITAR ALUMNO", command = lambda: self.editStudentUI(xlObj2, nameSelected, nameSelectedFlag)); self.editarAlumnoButton.grid(row = 22, column = 6);
        self.bajaAlumnoButton = tk.Button(self.findStudent, text = "BAJA DE ALUMNO", command = lambda: self.dropOutStudent(xlObj2, nameSelected, nameSelectedFlag)); self.bajaAlumnoButton.grid(row = 22, column = 7);

    def editStudentUI(self, xlEditObject, nameSelected, nameSelectedFlag):
        self.editStudent = Toplevel(self.mainMenu)
        self.editStudent.iconbitmap("imgs/Papalote.ico")
        self.editStudent.title("Editar Alumno")
        self.editStudent.geometry("800x200")

        #Labels
        self.nombreEditLabel = tk.Label(self.editStudent, text = "Nombre:"); self.nombreEditLabel.grid(column =0, row = 0);
        self.fechaNacimientoEditLabel = tk.Label(self.editStudent, text = "Fecha de Nacimiento:"); self.fechaNacimientoEditLabel.grid(column = 0, row = 1)
        self.curpEditLabel = tk.Label(self.editStudent, text = "CURP:"); self.curpEditLabel.grid(column = 0, row = 2)
        self.telefonoEditLabel = tk.Label(self.editStudent, text = "Telefono"); self.telefonoEditLabel.grid(column = 0, row = 3)
        self.otroTelefonoEditLabel = tk.Label(self.editStudent, text = "Otro Telefono:"); self.otroTelefonoEditLabel.grid(column = 0, row = 4)

        #PoT
        self.nombrePoTEditLabel = tk.Label(self.editStudent, text = "Nombre Padre o Tutor:"); self.nombrePoTEditLabel.grid(column = 2, row = 0)
        self.curpPoTEditLabel = tk.Label(self.editStudent, text="CURP:");self.curpPoTEditLabel.grid(column=2, row=1)
        self.rfcPoTEditLabel = tk.Label(self.editStudent, text="R.F.C.:");self.rfcPoTEditLabel.grid(column=2, row=2)
        self.telefonoPoTEditLabel = tk.Label(self.editStudent, text="Telefono:");self.telefonoPoTEditLabel.grid(column=2, row=3)
        self.celularPoTEditLabel = tk.Label(self.editStudent, text="Celular:");self.celularPoTEditLabel.grid(column=2, row=4)
        self.emailPoTEditLabel = tk.Label(self.editStudent, text="Email:"); self.emailPoTEditLabel.grid(column=2, row=5)

        #MoT
        self.nombreMoTEditLabel = tk.Label(self.editStudent, text="Nombre Padre o Tutor:");self.nombreMoTEditLabel.grid(column=4, row=0)
        self.curpMoTEditLabel = tk.Label(self.editStudent, text="CURP:");self.curpMoTEditLabel.grid(column=4, row=1)
        self.rfcMoTEditLabel = tk.Label(self.editStudent, text="R.F.C.:");self.rfcMoTEditLabel.grid(column=4, row=2)
        self.telefonoMoTEditLabel = tk.Label(self.editStudent, text="Telefono:");self.telefonoMoTEditLabel.grid(column=4, row=3)
        self.celularMoTEditLabel = tk.Label(self.editStudent, text="Celular:");self.celularMoTEditLabel.grid(column=4, row=4)
        self.emailMoTEditLabel = tk.Label(self.editStudent, text="Email:");self.emailMoTEditLabel.grid(column=4, row=5)


        #Entries
        self.nombreEditEntry = tk.Entry(self.editStudent); self.nombreEditEntry.grid(column = 1, row = 0)
        self.fechaNacimientoEditEntry =  tk.Entry(self.editStudent); self.fechaNacimientoEditEntry.grid(column = 1, row = 1)
        self.curpEditEntry = tk.Entry(self.editStudent); self.curpEditEntry.grid(column = 1, row = 2)
        self.telefonoEditEntry = tk.Entry(self.editStudent); self.telefonoEditEntry.grid(column = 1, row = 3)
        self.otroTelefonoEditEntry = tk.Entry(self.editStudent); self.otroTelefonoEditEntry.grid(column = 1, row = 4)

        #PoT
        self.nombrePoTEditEntry = tk.Entry(self.editStudent); self.nombrePoTEditEntry.grid(column = 3, row = 0);
        self.curpPoTEditEntry = tk.Entry(self.editStudent); self.curpPoTEditEntry.grid(column=3, row=1);
        self.rfcPoTEditEntry = tk.Entry(self.editStudent); self.rfcPoTEditEntry.grid(column=3, row=2);
        self.telefonoPoTEditEntry = tk.Entry(self.editStudent); self.telefonoPoTEditEntry.grid(column=3, row=3);
        self.celularPoTEditEntry = tk.Entry(self.editStudent); self.celularPoTEditEntry.grid(column=3, row=4);
        self.emailPoTEditEntry = tk.Entry(self.editStudent);self.emailPoTEditEntry.grid(column=3, row=5);

        #MoT
        self.nombreMoTEditEntry = tk.Entry(self.editStudent);self.nombreMoTEditEntry.grid(column=5, row=0);
        self.curpMoTEditEntry = tk.Entry(self.editStudent);self.curpMoTEditEntry.grid(column=5, row=1);
        self.rfcMoTEditEntry = tk.Entry(self.editStudent);self.rfcMoTEditEntry.grid(column=5, row=2);
        self.telefonoMoTEditEntry = tk.Entry(self.editStudent);self.telefonoMoTEditEntry.grid(column=5, row=3);
        self.celularMoTEditEntry = tk.Entry(self.editStudent);self.celularMoTEditEntry.grid(column=5, row=4);
        self.emailMoTEditEntry = tk.Entry(self.editStudent);self.emailMoTEditEntry.grid(column=5, row=5);

        #Buttons
        self.guardarEditButton = tk.Button(self.editStudent, text="GUARDAR", command=lambda: self.setEditedStudentData(xlEditObject));self.guardarEditButton.grid(row=6, column=5, sticky=E)

        #get Data
        self.getDataToEditStudent(xlEditObject, nameSelected, nameSelectedFlag)


    def updateStudentUI(self):
        if not self.updateStudent:
            self.updateStudent = Toplevel(self.mainMenu)
            self.updateStudent.iconbitmap("imgs/Papalote.ico")
            self.updateStudent.title("Actualizar Alumnos De Año")
            self.updateStudent.geometry("560x100")

        self.updateStudent.protocol("WM_DELETE_WINDOW", self.destroyUpdateStudent)

        xlObj3 = rousseauXL()
        #Labels
        self.warningLabel = tk.Label(self.updateStudent, text="AL DAR CLICK AL BOTON 'Actualizar Alumnos.xlsx', USTED ESTA DE ACUERDO EN HABER DADO DE BAJA \nA TODOS LOS ALUMNOS (EN LA SECCION BUSCAR ALUMNO O MANUALMENTE) QUE NO \nFUERON INSCRITOS A ESTE NUEVO CICLO ESCOLAR. YA QUE DE NO SER ASI LA BASE DE DATOS \nSE PODRIA VER AFECTADA Y SI NO HAY UN RESPALDO DE ELLA, LOS DATOS SERAN IRRECUPERABLES.");
        self.warningLabel.grid(column=0, row=0, sticky=N+S+E+W, columnspan = 2);

        #Buttons
        self.backupButton = tk.Button(self.updateStudent, text = "Respaldar Alumnos.xlsx", command= self.createBackUp).grid(column = 0, row = 1)
        self.actualizarAlumnosButton = tk.Button(self.updateStudent, text = "Actualizar Alumnos.xlsx" ,command= lambda: self.updateStudents(xlObj3)).grid(column = 1, row = 1)




    #Destroy window functions
    def destroyFindStudent(self):
        self.findStudent.destroy()
        self.findStudent = None

    def destroyUpdateStudent(self):
        self.updateStudent.destroy()
        self.updateStudent = None


    # Logic Functions
    def getDataFromNewStudent(self):

        self.progressBar["value"] = 10
        newStudentDict = {
            "nombre": self.nombreEntry.get(),
            "fechaNacimiento": self.fechaNacimientoEntry.get(),
            "curp": self.curpEntry.get(),
            "gradoACursar": self.gradoACursarVar.get(),
            "cicloEscolar": self.cicloEscolarVar.get(),
            "escuelaProcedencia": self.escuelaProcedenciaEntry.get(),
            "clave": self.claveEntry.get(),
            "conQuienVive": self.responsabeRBVar.get(),
            "calle": self.calleEntry.get(),
            "colonia": self.coloniaEntry.get(),
            "entreCalles": self.entreCallesEntry.get(),
            "codigoPostal": self.cpEntry.get(),
            "ciudad": self.ciudadEntry.get(),
            "telefono": self.telEntry.get(),
            "telefono2": self.tel2Entry.get(),
            "religion": self.religionEntry.get(),
            "enfermedadesOAccidentes": self.enfermedadesEntry.get(),
            "tipoSangre": self.sangreEntry.get(),
            "actualmenteTratamiento": self.sinoRBVar.get(),
            "servicioMedico1": self.imssCBVar.get(),
            "servicioMedico2": self.isssteCBVar.get(),
            "servicioMedico3": self.pemexCBVar.get(),
            "servicioMedico4": self.otroEntry.get()
        }

        newStudentPoTDict = {
            "nombrePoT": self.nombrePoTEntry.get(),
            "fechaNacimientoPoT": self.fechaNacimientoPoTEntry.get(),
            "curpPoT": self.curpPoTEntry.get(),
            "rfcPoT": self.rfcPoTEntry.get(),
            "lugarNacimientoPoT": self.lugarNacimientoPoTEntry.get(),
            "estadoCivilPoT": self.estadoCivilPoTEntry.get(),
            "nacionalidadPoT": self.nacionalidadPoTEntry.get(),
            "profesionPoT": self.profesionPoTEntry.get(),
            "telefonoPoT": self.telefonoPoTEntry.get(),
            "celularPoT": self.celularPoTEntry.get(),
            "lugarTrabajoPoT": self.lugarTrabajoPoTEntry.get(),
            "ocupacionPoT": self.ocupacionPoTEntry.get(),
            "emailPoT": self.emailPoTEntry.get()
        }

        newStudentMoTDict = {
            "nombreMoT": self.nombreMoTEntry.get(),
            "fechaNacimientoMoT": self.fechaNacimientoMoTEntry.get(),
            "curpMoT": self.curpMoTEntry.get(),
            "rfcMoT": self.rfcMoTEntry.get(),
            "lugarNacimientoMoT": self.lugarNacimientoMoTEntry.get(),
            "estadoCivilMoT": self.estadoCivilMoTEntry.get(),
            "nacionalidadMoT": self.nacionalidadMoTEntry.get(),
            "profesionMoT": self.profesionMoTEntry.get(),
            "telefonoMoT": self.telefonoMoTEntry.get(),
            "celularMoT": self.celularMoTEntry.get(),
            "lugarTrabajoMoT": self.lugarTrabajoMoTEntry.get(),
            "ocupacionMoT": self.ocupacionMoTEntry.get(),
            "emailMoT": self.emailMoTEntry.get()
        }

        newStudentReference = {
            "referencia": self.referenciaRBVar.get()
        }

        print(newStudentDict)
        print(newStudentPoTDict)
        print(newStudentMoTDict)
        siguienteAno = self.siguienteAnoCBVar.get()


        if newStudentDict["nombre"] == "":
            self.showTextBox("Error", "FAVOR DE INTRODUCIR EL NOMBRE DEL ALUMNO")
            del newStudentDict
            del newStudentPoTDict
            del newStudentMoTDict
            del newStudentReference
        elif newStudentDict["conQuienVive"] == "0":
            self.showTextBox("Error", "FAVOR DE INTRODUCIR CON QUIEN VIVE EL ALUMNO")
            del newStudentDict
            del newStudentPoTDict
            del newStudentMoTDict
            del newStudentReference
        elif newStudentDict["actualmenteTratamiento"] == "0":
            self.showTextBox("Error", "FAVOR DE INTRODUCIR SI EL ALUMNO SE ENCUENTRA EN TRATAMIENTO")
            del newStudentDict
            del newStudentPoTDict
            del newStudentMoTDict
            del newStudentReference
        elif siguienteAno:
            print("Alumno agregado para el siguiente ciclo escolar")
            try:
                xlObj = rousseauXL()
                self.progressBar["value"] = 60
                self.alumnosxlsxFlag = True
            except:
                messagebox.showerror("ERROR", "NO SE ENCONTRO 'Alummnos.xlsx'")
                self.newStudent.destroy()
                self.alumnosxlsxFlag = False


            if self.alumnosxlsxFlag:
                xlObj.moveAndWriteToNextYearSheet()
                xlObj.findRowToWrite()
                xlObj.addNewStudent(newStudentDict, newStudentPoTDict, newStudentMoTDict, newStudentReference)
                self.progressBar["value"] = 80
                self.clearDataFromNewStudent()
                xlObj.moveBackToAlumnosSheet()
                xlObj.save()
                self.progressBar["value"] = 100
                del self.alumnosxlsxFlag
                del xlObj
                self.progressBar["value"] = 0
                self.showTextBox("Info", "Alumno(A) {} GUARDADO CON EXITO PARA EL SIGUIENTE CICLO!".format(newStudentDict["nombre"]))

            return
        else:
            self.progressBar["value"] = 40
            try:
                xlObj = rousseauXL()
                self.progressBar["value"] = 60
                self.alumnosxlsxFlag = True
            except:
                messagebox.showerror("ERROR", "NO SE ENCONTRO 'Alummnos.xlsx'")
                self.newStudent.destroy()
                self.alumnosxlsxFlag = False

            if self.alumnosxlsxFlag:
                if xlObj.validateSheet():
                    self.progressBar["value"] = 80
                    xlObj.findRowToWrite()
                    self.progressBar["value"] = 90
                    xlObj.addNewStudent(newStudentDict, newStudentPoTDict, newStudentMoTDict, newStudentReference)
                    self.progressBar["value"] = 95
                    self.clearDataFromNewStudent()
                    xlObj.save()
                    self.progressBar["value"] = 100
                    self.showTextBox("Info", "ALUMNO(A) {} GUARDADO CON EXITO!".format(newStudentDict["nombre"]))
                    del xlObj
                    self.progressBar["value"] = 0
                    pass
                else:
                    self.progressBar["value"] = 0
                    self.showTextBox("Error", "ALUMNO NO AGREGADO, VERIFIQUE EL ARCHIVO DE EXCEL 'Alumnos.xlsx'")

    def clearDataFromNewStudent(self):
        self.nombreEntry.delete(0, 'end');
        self.fechaNacimientoEntry.delete(0, 'end');
        self.curpEntry.delete(0, 'end');
        self.escuelaProcedenciaEntry.delete(0, 'end');
        self.claveEntry.delete(0, 'end');
        self.calleEntry.delete(0, 'end');
        self.coloniaEntry.delete(0, 'end');
        self.entreCallesEntry.delete(0, 'end');
        self.cpEntry.delete(0, 'end');
        self.ciudadEntry.delete(0, 'end');
        self.telEntry.delete(0, 'end');
        self.tel2Entry.delete(0, 'end');
        self.religionEntry.delete(0, 'end');
        self.enfermedadesEntry.delete(0, 'end');
        self.sangreEntry.delete(0, 'end');
        self.otroEntry.delete(0, 'end');
        self.nombrePoTEntry.delete(0, 'end');
        self.fechaNacimientoPoTEntry.delete(0, 'end');
        self.curpPoTEntry.delete(0, 'end');
        self.rfcPoTEntry.delete(0, 'end');
        self.lugarNacimientoPoTEntry.delete(0, 'end');
        self.estadoCivilPoTEntry.delete(0, 'end');
        self.nacionalidadPoTEntry.delete(0, 'end');
        self.profesionPoTEntry.delete(0, 'end');
        self.telefonoPoTEntry.delete(0, 'end');
        self.celularPoTEntry.delete(0, 'end');
        self.lugarTrabajoPoTEntry.delete(0, 'end');
        self.ocupacionPoTEntry.delete(0, 'end');
        self.emailPoTEntry.delete(0, 'end');
        self.nombreMoTEntry.delete(0, 'end');
        self.fechaNacimientoMoTEntry.delete(0, 'end');
        self.curpMoTEntry.delete(0, 'end');
        self.rfcMoTEntry.delete(0, 'end');
        self.lugarNacimientoMoTEntry.delete(0, 'end');
        self.estadoCivilMoTEntry.delete(0, 'end');
        self.nacionalidadMoTEntry.delete(0, 'end');
        self.profesionMoTEntry.delete(0, 'end');
        self.telefonoMoTEntry.delete(0, 'end');
        self.celularMoTEntry.delete(0, 'end');
        self.lugarTrabajoMoTEntry.delete(0, 'end');
        self.ocupacionMoTEntry.delete(0, 'end');
        self.emailMoTEntry.delete(0, 'end');

        self.imssCBVar.set(False);
        self.isssteCBVar.set(False);
        self.pemexCBVar.set(False);
        self.sinoRBVar.set(False);
        self.referenciaRBVar.set(False);
        self.responsabeRBVar.set(False);
        self.siguienteAnoCBVar.set(False);

    def getAndVerifyUsernameAndPassword(self):
        # Available usernames and passwords
        validUsernames = "1"
        validPasswords = "1"
        username = self.usernameEntry.get()
        password = self.passwordEntry.get()

        if validUsernames == username and validPasswords == password:
            self.master.withdraw()
            self.mainMenuUI()
        else:
            self.showTextBox("Error", "Verifique su usuario o contrasena")

    def showTextBox(self, typeOfMessage, message):

        if typeOfMessage == "Info":
            messagebox.showinfo("Info", message)
        elif typeOfMessage == "Warning":
            messagebox.showwarning("Warning", message)
        elif typeOfMessage == "Error":
            messagebox.showerror("Error", message)

    def getDataFromSpecificStudent(self, xlObject, nameSelected, nameSelectedFlag, updateLabels):
        self.nombreAlumnoResultLabel.config(text = "-")
        self.fechaNacimientoAlumnoResultLabel.config(text = "-")
        self.curpAlumnoResultLabel.config(text = "-")
        self.gradoACursarResultLabel.config(text = "-")
        self.cicloEscolarResultLabel.config(text = "-")
        self.gradoEnCursoResultLabel.config(text = "-")
        self.cicloEnCursoResultLabel.config(text = "-")
        self.conQuienViveResultLabel.config(text = "-")
        self.calleAlumnoResultLabel.config(text = "-")
        self.coloniaAlumnoResultLabel.config(text = "-")
        self.cpAlumnoResultLabel.config(text = "-")
        self.ciudadAlumnoResultLabel.config(text = "-")
        self.entreCallesAlumnoResultLabel.config(text = "-")
        self.telefonoAlumnoResultLabel.config(text = "-")
        self.telefono2AlumnoResultLabel.config(text = "-")
        self.enfermedadesResultLabel.config(text = "-")
        self.tipoSangreResultLabel.config(text = "-")
        self.enTratamientoResultLabel.config(text = "-")
        self.servicioMedicoResultLabel.config(text = "-")
        self.fechaInscripcionResultLabel.config(text = "-")

        self.nombrePotResultLabel.config(text = "-")
        self.curpPotResultLabel.config(text = "-")
        self.rfcPotResultLabel.config(text = "-")
        self.estadoCivilPotResultLabel.config(text = "-")
        self.nacionalidadPotResultLabel.config(text = "-")
        self.profesionPotResultLabel.config(text = "-")
        self.ocupacionPotResultLabel.config(text = "-")
        self.lugarDeTrabajoPotResultLabel.config(text = "-")
        self.telefonoPotResultLabel.config(text = "-")
        self.celularPotResultLabel.config(text = "-")
        self.emailPotResultLabel.config(text = "-")

        self.nombreMotResultLabel.config(text = "-")
        self.curpMotResultLabel.config(text = "-")
        self.rfcMotResultLabel.config(text = "-")
        self.estadoCivilMotResultLabel.config(text = "-")
        self.nacionalidadMotResultLabel.config(text = "-")
        self.profesionMotResultLabel.config(text = "-")
        self.ocupacionMotResultLabel.config(text = "-")
        self.lugarDeTrabajoMotResultLabel.config(text = "-")
        self.telefonoMotResultLabel.config(text = "-")
        self.celularMotResultLabel.config(text = "-")
        self.emailMotResultLabel.config(text = "-")

        self.statusResultLabel.config(text="-")

        if nameSelectedFlag:
            c, r = xlObject.getRowFromSpecificStudent(nameSelected)
            #print(str(c) + " " + str(r))
            self.dataDict = xlObject.getToShowAllDataFromSpecificStudent(c, r)
            print(self.dataDict)

            #Update Labels
            self.nombreAlumnoResultLabel.config(text = self.dataDict["nombre"])
            self.fechaNacimientoAlumnoResultLabel.config(text = self.dataDict["fechaNacimiento"])
            self.curpAlumnoResultLabel.config(text = self.dataDict["curp"])
            self.gradoACursarResultLabel.config(text = self.dataDict["gradoACursar"])
            self.cicloEscolarResultLabel.config(text = self.dataDict["cicloEscolar"])
            self.gradoEnCursoResultLabel.config(text = self.dataDict["gradoEnCurso"])
            self.cicloEnCursoResultLabel.config(text = self.dataDict["cicloEscolarEnCurso"])
            self.conQuienViveResultLabel.config(text = self.dataDict["conQuienVive"])
            self.calleAlumnoResultLabel.config(text = self.dataDict["calle"])
            self.coloniaAlumnoResultLabel.config(text = self.dataDict["colonia"])
            self.cpAlumnoResultLabel.config(text = self.dataDict["codigoPostal"])
            self.ciudadAlumnoResultLabel.config(text = self.dataDict["ciudad"])
            self.entreCallesAlumnoResultLabel.config(text = self.dataDict["entreCalles"])
            self.telefonoAlumnoResultLabel.config(text = self.dataDict["telefono"])
            self.telefono2AlumnoResultLabel.config(text = self.dataDict["telefono2"])
            self.enfermedadesResultLabel.config(text = self.dataDict["enfermedadesOAccidentes"])
            self.tipoSangreResultLabel.config(text = self.dataDict["tipoSangre"])
            self.enTratamientoResultLabel.config(text = self.dataDict["actualmenteTratamiento"])
            self.servicioMedicoResultLabel.config(text = self.dataDict["servicioMedico1"])
            self.fechaInscripcionResultLabel.config(text = self.dataDict["fechaInscripcion"])

            self.nombrePotResultLabel.config(text = self.dataDict["nombrePoT"])
            self.curpPotResultLabel.config(text = self.dataDict["curpPoT"])
            self.rfcPotResultLabel.config(text = self.dataDict["rfcPoT"])
            self.estadoCivilPotResultLabel.config(text = self.dataDict["estadoCivilPoT"])
            self.nacionalidadPotResultLabel.config(text = self.dataDict["nacionalidadPoT"])
            self.profesionPotResultLabel.config(text = self.dataDict["profesionPoT"])
            self.ocupacionPotResultLabel.config(text = self.dataDict["ocupacionPoT"])
            self.lugarDeTrabajoPotResultLabel.config(text = self.dataDict["lugarTrabajoPoT"])
            self.telefonoPotResultLabel.config(text = self.dataDict["telefonoPoT"])
            self.celularPotResultLabel.config(text = self.dataDict["celularPoT"])
            self.emailPotResultLabel.config(text = self.dataDict["emailPoT"])

            self.nombreMotResultLabel.config(text = self.dataDict["nombreMoT"])
            self.curpMotResultLabel.config(text = self.dataDict["curpMoT"])
            self.rfcMotResultLabel.config(text = self.dataDict["rfcMoT"])
            self.estadoCivilMotResultLabel.config(text = self.dataDict["estadoCivilMoT"])
            self.nacionalidadMotResultLabel.config(text = self.dataDict["nacionalidadMoT"])
            self.profesionMotResultLabel.config(text = self.dataDict["profesionMoT"])
            self.ocupacionMotResultLabel.config(text = self.dataDict["ocupacionMoT"])
            self.lugarDeTrabajoMotResultLabel.config(text = self.dataDict["lugarTrabajoMoT"])
            self.telefonoMotResultLabel.config(text = self.dataDict["telefonoMoT"])
            self.celularMotResultLabel.config(text = self.dataDict["celularMoT"])
            self.emailMotResultLabel.config(text = self.dataDict["emailMoT"])

            self.statusResultLabel.config(text = self.dataDict["status"])


        elif nameSelectedFlag == False and updateLabels == True:
            return
        else:
            self.showTextBox("Error", "SELECCIONE AL ALUMNO QUE DESEA BUSCAR")

    def deleteSpecificStudent(self, xlDeleteObject, nameSelected, nameSelectedFlag):
            if nameSelectedFlag:
                try:
                    c, r = xlDeleteObject.getRowFromSpecificStudent(nameSelected)
                except:
                    self.showTextBox("Error", "SELECCIONE AL ALUMNO QUE DESEA BORRAR")

                xlDeleteObject.deleteSpecificStudent(r)
                xlDeleteObject.save()
                newList = xlDeleteObject.getNamesList()

                #Update List to avoid showing the deleted student
                self.nombresListbox.delete(0, 'end')
                data = sorted(newList, key=str.lower)
                for item in data:
                    self.nombresListbox.insert('end', item)

                #Delete student from labels
                nameSelectedFlag = False
                self.getDataFromSpecificStudent(xlDeleteObject, nameSelected, nameSelectedFlag, True)
                del xlDeleteObject
                self.showTextBox("Info", "EL ALUMNO {} FUE BORRADO CON EXITO".format(nameSelected))

            else:
                self.showTextBox("Error", "SELECCIONE AL ALUMNO QUE DESEA BORRAR")

    def dropOutStudent(self, xlDropOutObject, nameSelected, nameSelectedFlag):
        if nameSelectedFlag:
            try:
                c, r = xlDropOutObject.getRowFromSpecificStudent(nameSelected)
            except:
                self.showTextBox("Error", "SELECCIONE AL ALUMNO QUE DESEA DAR DE BAJA")

            xlDropOutObject.dropOutStudent(r)
            xlDropOutObject.save()
            del xlDropOutObject
            self.showTextBox("Info", "EL ALUMNO {} FUE DADO DE BAJA CON EXITO".format(nameSelected))
        else:
            self.showTextBox("Error", "SELECCIONE AL ALUMNO QUE DESEA DAR DE BAJA")

    def getDataToEditStudent(self, xlEditStudentObject, nameSelected, nameSelectedFlag):
        if nameSelected:
            c, r = xlEditStudentObject.getRowFromSpecificStudent(nameSelected)

            self.rowToEditedStudent = r
            editDict = xlEditStudentObject.getToShowAllDataFromSpecificStudent(c, r)


            #To avoid empty elements
            for elements, data in editDict.items():
                if data == None:
                    editDict[elements] = " "

            self.nombreEditEntry.insert(0, editDict["nombre"])
            self.fechaNacimientoEditEntry.insert(0, editDict["fechaNacimiento"])
            self.curpEditEntry.insert(0, editDict["curp"])
            self.telefonoEditEntry.insert(0, editDict["telefono"])
            self.otroTelefonoEditEntry.insert(0, editDict["telefono2"])

            #PoT
            self.nombrePoTEditEntry.insert(0,editDict["nombrePoT"])
            self.curpPoTEditEntry.insert(0,editDict["curpPoT"])
            self.rfcPoTEditEntry.insert(0,editDict["rfcPoT"])
            self.telefonoPoTEditEntry.insert(0,editDict["telefonoPoT"])
            self.celularPoTEditEntry.insert(0,editDict["celularPoT"])
            self.emailPoTEditEntry.insert(0,editDict["emailPoT"])

            #MoT
            self.nombreMoTEditEntry.insert(0, editDict["nombreMoT"])
            self.curpMoTEditEntry.insert(0, editDict["curpMoT"])
            self.rfcMoTEditEntry.insert(0, editDict["rfcMoT"])
            self.telefonoMoTEditEntry.insert(0, editDict["telefonoMoT"])
            self.celularMoTEditEntry.insert(0, editDict["celularMoT"])
            self.emailMoTEditEntry.insert(0, editDict["emailMoT"])


        else:
            self.showTextBox("Error", "SELECCIONE AL ALUMNO(A) QUE DESEA EDITAR")
            self.editStudent.destroy()

    def setEditedStudentData(self, xlEditedStudentObject):
        editedDict = {
            "editedNombre": self.nombreEditEntry.get(),
            "editedFechaNacimiento": self.fechaNacimientoEditEntry.get(),
            "editedCurp": self.curpEditEntry.get(),
            "editedTelefono": self.telefonoEditEntry.get(),
            "editedOtroTelefono": self.otroTelefonoEditEntry.get(),

            "editedNombrePoT": self.nombrePoTEditEntry.get(),
            "editedCurpPoT": self.curpPoTEditEntry.get(),
            "editedRFCPoT": self.rfcPoTEditEntry.get(),
            "editedTelPoT": self.telefonoPoTEditEntry.get(),
            "editedCelPoT": self.celularPoTEditEntry.get(),
            "editedEmailPoT": self.emailPoTEditEntry.get(),

            "editedNombreMoT": self.nombreMoTEditEntry.get(),
            "editedCurpMoT": self.curpMoTEditEntry.get(),
            "editedRFCMoT": self.rfcMoTEditEntry.get(),
            "editedTelMoT": self.telefonoMoTEditEntry.get(),
            "editedCelMoT": self.celularMoTEditEntry.get(),
            "editedEmailMoT": self.emailMoTEditEntry.get()
        }
        xlEditedStudentObject.setEditedStudentData(editedDict, self.rowToEditedStudent)
        xlEditedStudentObject.save()
        del xlEditedStudentObject

        self.showTextBox("Info", "ALUMNO EDITADO EXITOSAMENTE")
        self.editStudent.destroy()
        self.destroyFindStudent()


    def updateStudents(self, xlUpdateStudentsObject):

        warningMsgBox = tk.messagebox.askyesno("Verifique", "ESTA SEGURO(A) DE PROCEDER CON ESTA OPERACION?.\nSI RESPONDE QUE SI, ESTO PODRIA DEMORAR UNOS MINUTOS \nAL FINALIZAR SERA NOTIFICADO(A), FAVOR DE NO CERRAR LA VENTANA.",icon = 'warning')
        print(warningMsgBox)
        if warningMsgBox == True:
            print("yes")
            xlUpdateStudentsObject.updateStudent()
            xlUpdateStudentsObject.save()
            self.showTextBox("Info", "LA BASE DE DATOS: 'Alumnos.xlsx' FUE ACTUALIZADA CON EXITO")
        else:
            pass
        del xlUpdateStudentsObject

    def createBackUp(self):
        ruta = os.getcwd()
        time = datetime.datetime.now().strftime("%Y-%m-%d")
        src_dir = "Alumnos.xlsx"
        dst_dir = "./Respaldo/Alumnos_"+str(time)+".xlsx"

        if os.path.isdir('./Respaldo') == True:
            shutil.copy(src_dir,dst_dir)
            self.showTextBox("Info", "Respaldo creado en la carpeta 'Respaldo'.")
        else:
            access_rights = 0o777
            try:
                os.mkdir(ruta+"\Respaldo", access_rights)
                shutil.copy(src_dir, dst_dir)
                self.showTextBox("Info", "Respaldo creado en la carpeta 'Respaldo'.")
            except OSError:
                self.showTextBox("Info","NO SE PUDO CREAR RESPALDO.")




class rousseauXL:
    global wb
    global ws
    global emptyRow
    global emptyCol

    def __init__(self):
        self.wb = xl.load_workbook(filename="Alumnos.xlsx")
        self.ws = self.wb['Alumnos']

    def validateSheet(self):
        nombreCellValidation = str(self.ws.cell(row=1, column=1).value)
        if nombreCellValidation == "NOMBRE":
            nombreFlag = True
        else:
            nombreFlag = False
        return nombreFlag

    def moveAndWriteToNextYearSheet(self):
        try:
            self.ws = self.wb["Siguiente ciclo"]
        except:
            self.wb.create_sheet('Siguiente ciclo')
            self.ws = self.wb['Siguiente ciclo']


    def moveBackToAlumnosSheet(self):
        self.ws = self.wb['Alumnos']

    def findRowToWrite(self):
        for row in range(2, self.ws.max_row + 2):
            for column in "A":
                cell_name = "{}{}".format(column, row)
                # print("{},{}".format(column, row) + str(self.ws[cell_name].value))
                if str(self.ws[cell_name].value) == "None":
                    # Obtener la celda en la columna A "NOMBRE" donde no hay nada escrito para insertar ahi el nuevo alumno
                    self.emptyRow = row
                    self.emptyCol = column

    def addNewStudent(self, newStudentDict, newStudentPoTDict, newStudentMoTDict, newStudentReference):
        #done
        print(self.emptyCol, self.emptyRow)
        self.ws.cell(row=self.emptyRow, column=1).value = newStudentDict["nombre"]
        self.ws.cell(row=self.emptyRow, column=2).value = newStudentDict["fechaNacimiento"]
        self.ws.cell(row=self.emptyRow, column=3).value = newStudentDict["curp"]
        self.ws.cell(row=self.emptyRow, column=4).value = newStudentDict["gradoACursar"]
        self.ws.cell(row=self.emptyRow, column=5).value = newStudentDict["cicloEscolar"]
        self.ws.cell(row=self.emptyRow, column =6).value = newStudentDict["gradoACursar"]
        self.ws.cell(row=self.emptyRow, column=7).value = newStudentDict["cicloEscolar"]
        self.ws.cell(row=self.emptyRow, column=8).value = newStudentDict["escuelaProcedencia"]
        self.ws.cell(row=self.emptyRow, column=9).value = newStudentDict["clave"]
        self.ws.cell(row=self.emptyRow, column=10).value = newStudentDict["conQuienVive"]
        self.ws.cell(row=self.emptyRow, column=11).value = newStudentDict["calle"]
        self.ws.cell(row=self.emptyRow, column=12).value = newStudentDict["colonia"]
        self.ws.cell(row=self.emptyRow, column=13).value = newStudentDict["entreCalles"]
        self.ws.cell(row=self.emptyRow, column=14).value = newStudentDict["codigoPostal"]
        self.ws.cell(row=self.emptyRow, column=15).value = newStudentDict["ciudad"]
        self.ws.cell(row=self.emptyRow, column=16).value = newStudentDict["telefono"]
        self.ws.cell(row=self.emptyRow, column=17).value = newStudentDict["telefono2"]
        self.ws.cell(row=self.emptyRow, column=18).value = newStudentDict["religion"]
        self.ws.cell(row=self.emptyRow, column=19).value = newStudentDict["enfermedadesOAccidentes"]
        self.ws.cell(row=self.emptyRow, column=20).value = newStudentDict["tipoSangre"]
        self.ws.cell(row=self.emptyRow, column=21).value = newStudentDict["actualmenteTratamiento"]

        if newStudentDict["servicioMedico1"] == True:
            imss = "IMSS"
        else:
            imss = ""

        if newStudentDict["servicioMedico2"] == True:
            issste = "ISSSTE"
        else:
            issste = ""

        if newStudentDict["servicioMedico3"] == True:
            pemex = "PEMEX"
        else:
            pemex = ""
        servicioMedicoStr = imss + " " + issste + " " + pemex + " " + str(newStudentDict["servicioMedico4"])
        self.ws.cell(row=self.emptyRow, column=22).value = servicioMedicoStr

        # PoT
        self.ws.cell(row=self.emptyRow, column=23).value = newStudentPoTDict["nombrePoT"]
        self.ws.cell(row=self.emptyRow, column=24).value = newStudentPoTDict["fechaNacimientoPoT"]
        self.ws.cell(row=self.emptyRow, column=25).value = newStudentPoTDict["curpPoT"]
        self.ws.cell(row=self.emptyRow, column=26).value = newStudentPoTDict["rfcPoT"]
        self.ws.cell(row=self.emptyRow, column=27).value = newStudentPoTDict["lugarNacimientoPoT"]
        self.ws.cell(row=self.emptyRow, column=28).value = newStudentPoTDict["estadoCivilPoT"]
        self.ws.cell(row=self.emptyRow, column=29).value = newStudentPoTDict["nacionalidadPoT"]
        self.ws.cell(row=self.emptyRow, column=30).value = newStudentPoTDict["profesionPoT"]
        self.ws.cell(row=self.emptyRow, column=31).value = newStudentPoTDict["telefonoPoT"]
        self.ws.cell(row=self.emptyRow, column=32).value = newStudentPoTDict["celularPoT"]
        self.ws.cell(row=self.emptyRow, column=33).value = newStudentPoTDict["lugarTrabajoPoT"]
        self.ws.cell(row=self.emptyRow, column=34).value = newStudentPoTDict["ocupacionPoT"]
        self.ws.cell(row=self.emptyRow, column=35).value = newStudentPoTDict["emailPoT"]

        # MoT
        self.ws.cell(row=self.emptyRow, column=36).value = newStudentMoTDict["nombreMoT"]
        self.ws.cell(row=self.emptyRow, column=37).value = newStudentMoTDict["fechaNacimientoMoT"]
        self.ws.cell(row=self.emptyRow, column=38).value = newStudentMoTDict["curpMoT"]
        self.ws.cell(row=self.emptyRow, column=39).value = newStudentMoTDict["rfcMoT"]
        self.ws.cell(row=self.emptyRow, column=40).value = newStudentMoTDict["lugarNacimientoMoT"]
        self.ws.cell(row=self.emptyRow, column=41).value = newStudentMoTDict["estadoCivilMoT"]
        self.ws.cell(row=self.emptyRow, column=42).value = newStudentMoTDict["nacionalidadMoT"]
        self.ws.cell(row=self.emptyRow, column=43).value = newStudentMoTDict["profesionMoT"]
        self.ws.cell(row=self.emptyRow, column=44).value = newStudentMoTDict["telefonoMoT"]
        self.ws.cell(row=self.emptyRow, column=45).value = newStudentMoTDict["celularMoT"]
        self.ws.cell(row=self.emptyRow, column=46).value = newStudentMoTDict["lugarTrabajoMoT"]
        self.ws.cell(row=self.emptyRow, column=47).value = newStudentMoTDict["ocupacionMoT"]
        self.ws.cell(row=self.emptyRow, column=48).value = newStudentMoTDict["emailMoT"]

        self.ws.cell(row=self.emptyRow, column=49).value = newStudentReference["referencia"]
        self.ws.cell(row=self.emptyRow, column=50).value = datetime.datetime.now().strftime("%Y-%m-%d")

        # for column in range(1, self.ws.max_column+1):
        #     self.ws.cell(row = self.emptyRow, column = column).value = "si"

    def save(self):
        self.wb.save(filename="Alumnos.xlsx")

    def getNamesList(self):
        names_list = []
        for row in range(2, self.ws.max_row + 1):
            for column in "A":
                cell_name = "{}{}".format(column, row)
                # print("ROW: " + str(row) )
                # print("{},{}".format(column, row) + str(ws[cell_name].value))
                names_list.append(str(self.ws[cell_name].value))

        return names_list

    def getRowFromSpecificStudent(self, specificName):
        #print("Buscando a : " + str(specificName))
        for row in range(2,self.ws.max_row + 1):
            for column in "A":
                cell_name = "{}{}".format(column, row)
                print(cell_name)
                if specificName == str(self.ws[cell_name].value):
                    #print("::::::::" + "{},{}".format(column, row) + str(self.ws[cell_name].value))
                    return column, row

    def getToShowAllDataFromSpecificStudent(self, col, row):
        #done
        def dropedOrActive():
            if str(self.ws.cell(row=row, column=51).value) == "BAJA":
                return "DADO DE BAJA"
            else:
                return "ACTIVO"
        dict = {
            "nombre": self.ws.cell(row=row, column=1).value,
            "fechaNacimiento": self.ws.cell(row=row, column=2).value,
            "curp": self.ws.cell(row=row, column=3).value,
            "gradoACursar": self.ws.cell(row=row, column=4).value,
            "cicloEscolar": self.ws.cell(row=row, column=5).value,
            "gradoEnCurso": self.ws.cell(row=row, column = 6).value,
            "cicloEscolarEnCurso": self.ws.cell(row=row, column = 7).value,
            "conQuienVive": self.ws.cell(row=row, column=10).value,
            "calle": self.ws.cell(row=row, column=11).value,
            "colonia": self.ws.cell(row=row, column=12).value,
            "entreCalles": self.ws.cell(row=row, column=13).value,
            "codigoPostal": self.ws.cell(row=row, column=14).value,
            "ciudad": self.ws.cell(row=row, column=15).value,
            "telefono": self.ws.cell(row=row, column=16).value,
            "telefono2": self.ws.cell(row=row, column=17).value,
            "enfermedadesOAccidentes": self.ws.cell(row=row, column=19).value,
            "tipoSangre": self.ws.cell(row=row, column=20).value,
            "actualmenteTratamiento": self.ws.cell(row=row, column=21).value,
            "servicioMedico1": self.ws.cell(row=row, column=22).value,
            "fechaInscripcion": self.ws.cell(row = row , column = 50).value,

            "nombrePoT": self.ws.cell(row=row, column=23).value,
            "curpPoT": self.ws.cell(row=row, column=25).value,
            "rfcPoT": self.ws.cell(row=row, column=26).value,
            "estadoCivilPoT": self.ws.cell(row=row, column=28).value,
            "nacionalidadPoT": self.ws.cell(row=row, column=29).value,
            "profesionPoT": self.ws.cell(row=row, column=30).value,
            "telefonoPoT": self.ws.cell(row=row, column=31).value,
            "celularPoT": self.ws.cell(row=row, column=32).value,
            "lugarTrabajoPoT": self.ws.cell(row=row, column=33).value,
            "ocupacionPoT": self.ws.cell(row=row, column=34).value,
            "emailPoT": self.ws.cell(row=row, column=35).value,

            "nombreMoT": self.ws.cell(row=row, column=36).value,
            "curpMoT": self.ws.cell(row=row, column=38).value,
            "rfcMoT": self.ws.cell(row=row, column=39).value,
            "estadoCivilMoT": self.ws.cell(row=row, column=41).value,
            "nacionalidadMoT": self.ws.cell(row=row, column=42).value,
            "profesionMoT": self.ws.cell(row=row, column=43).value,
            "telefonoMoT": self.ws.cell(row=row, column=44).value,
            "celularMoT": self.ws.cell(row=row, column=45).value,
            "lugarTrabajoMoT": self.ws.cell(row=row, column=46).value,
            "ocupacionMoT": self.ws.cell(row=row, column=47).value,
            "emailMoT": self.ws.cell(row=row, column=48).value,

            "status": dropedOrActive()
        }
        return dict

    def setEditedStudentData(self, editedDict, row):
        self.ws.cell(row= row, column=1).value = editedDict["editedNombre"]
        self.ws.cell(row=row, column=2).value = editedDict["editedFechaNacimiento"]
        self.ws.cell(row=row, column=3).value = editedDict["editedCurp"]
        self.ws.cell(row=row, column=16).value = editedDict["editedTelefono"]
        self.ws.cell(row=row, column=17).value = editedDict["editedOtroTelefono"]

        self.ws.cell(row = row, column = 23).value = editedDict["editedNombrePoT"]
        self.ws.cell(row=row, column=25).value = editedDict["editedCurpPoT"]
        self.ws.cell(row=row, column=26).value = editedDict["editedRFCPoT"]
        self.ws.cell(row=row, column=31).value = editedDict["editedTelPoT"]
        self.ws.cell(row=row, column=32).value = editedDict["editedCelPoT"]
        self.ws.cell(row=row, column=35).value = editedDict["editedEmailPoT"]

        self.ws.cell(row=row, column=36).value = editedDict["editedNombreMoT"]
        self.ws.cell(row=row, column=38).value = editedDict["editedCurpMoT"]
        self.ws.cell(row=row, column=39).value = editedDict["editedRFCMoT"]
        self.ws.cell(row=row, column=44).value = editedDict["editedTelMoT"]
        self.ws.cell(row=row, column=45).value = editedDict["editedCelMoT"]
        self.ws.cell(row=row, column=48).value = editedDict["editedEmailMoT"]

    def deleteSpecificStudent(self, row):
        self.ws.delete_rows(row)

    def dropOutStudent(self, row):
        self.ws.cell(row=row, column=51).value = "BAJA"
        self.ws.cell(row=row, column=52).value = datetime.datetime.now().strftime("%Y-%m-%d")

        red_font = Font(color='00FF0000', italic=True)

        # Enumerate the cells in the second row
        for cell in self.ws[row:row]:
            cell.font = red_font

    def updateStudent(self):
        for row in range(2, self.ws.max_row + 1):
            currentYear = self.ws.cell(row=row, column=6).value
            currentYearValidation = currentYear
            if str(self.ws.cell(row=row, column=51).value) != "BAJA" and currentYear != "Se graduo de secundaria":
                currentCycle = self.ws.cell(row=row, column=7).value
                if "Kinder" in currentYear:
                    currentYear = currentYear.strip(" Kinder")
                    currentYear = int(currentYear)
                    if currentYear == 1:
                        newYear = "2 Kinder"
                    elif currentYear == 2:
                        newYear = "3 Kinder"
                    elif currentYear == 3:
                        newYear = "1 Primaria"

                elif "Primaria" in currentYear:
                    currentYear = currentYear.strip(" Primaria")
                    currentYear = int(currentYear)
                    if currentYear == 1:
                        newYear = "2 Primaria"
                    if currentYear == 2:
                        newYear = "3 Primaria"
                    if currentYear == 3:
                        newYear = "4 Primaria"
                    if currentYear == 4:
                        newYear = "5 Primaria"
                    if currentYear == 5:
                        newYear = "6 Primaria"
                    if currentYear == 6:
                        newYear = "1 Secundaria"

                elif "Secundaria" in currentYear:
                    currentYear = currentYear.strip(" Secundaria")
                    currentYear = int(currentYear)
                    if currentYear == 1:
                        newYear = "2 Secundaria"
                    elif currentYear == 2:
                        newYear = "3 Secundaria"
                    elif currentYear == 3:
                        newYear = "Se graduo de secundaria"
                elif "Se graduo de secundaria" in currentYear:
                    newYear = "Se graduo de secundaria"

                if currentYearValidation == "3 Secundaria":
                    newCycle = currentCycle
                if currentCycle is not "Se graduo en secundaria" and currentYearValidation != "3 Secundaria":
                    print(currentYearValidation)
                    first4DigitsCycle = int(currentCycle[:4])
                    last4DigitsCycle = int(currentCycle[5:])
                    first4DigitsCycle = first4DigitsCycle + 1
                    last4DigitsCycle = last4DigitsCycle + 1
                    newCycle = str(first4DigitsCycle) + "-" + str(last4DigitsCycle)

                self.ws.cell(row=row, column=6).value = newYear
                self.ws.cell(row=row, column=7).value = newCycle
                del currentYear
                del newYear
                del currentCycle
                try:
                    del first4DigitsCycle
                    del last4DigitsCycle
                except:
                    pass
                del newCycle

            else:
                pass

def main():
    root = tk.Tk()
    w = 250;
    h = 75;
    x = 50;
    y = 100
    root.geometry("%dx%d+%d+%d" % (w, h, x, y))
    root.iconbitmap("imgs/Papalote.ico")
    loginWindow = Rousseau(root)

    root.mainloop()


if __name__ == '__main__':
    main()
